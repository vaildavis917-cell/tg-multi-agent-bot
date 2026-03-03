"""
Парсинг файлов: PDF, Excel, CSV, TXT.
Извлекает текст для добавления в контекст агента.
"""

import os
import logging
from typing import Optional

logger = logging.getLogger(__name__)

MAX_FILE_TEXT_LENGTH = 15000  # Ограничение длины текста из файла


def parse_file(file_path: str) -> Optional[str]:
    """
    Парсит файл и возвращает текстовое содержимое.
    Поддерживает: .txt, .csv, .xlsx, .xls, .pdf
    """
    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".txt":
            return _parse_txt(file_path)
        elif ext == ".csv":
            return _parse_csv(file_path)
        elif ext in (".xlsx", ".xls"):
            return _parse_excel(file_path)
        elif ext == ".pdf":
            return _parse_pdf(file_path)
        else:
            logger.warning("Unsupported file type: %s", ext)
            return None
    except Exception as e:
        logger.error("Failed to parse %s: %s", file_path, e)
        return None


def _parse_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        text = f.read()
    return _truncate(text)


def _parse_csv(path: str) -> str:
    import csv
    lines = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i > 500:  # Ограничение строк
                lines.append(f"... (ещё строки, всего > 500)")
                break
            lines.append(" | ".join(row))
    return _truncate("\n".join(lines))


def _parse_excel(path: str) -> str:
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed")
        return None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Лист: {sheet_name} ===")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count > 500:
                lines.append("... (ещё строки)")
                break
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))
            row_count += 1

    wb.close()
    return _truncate("\n".join(lines))


def _parse_pdf(path: str) -> str:
    try:
        import subprocess
        result = subprocess.run(
            ["pdftotext", "-layout", path, "-"],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode == 0 and result.stdout.strip():
            return _truncate(result.stdout.strip())
    except (FileNotFoundError, subprocess.TimeoutExpired):
        pass

    # Fallback: PyPDF2
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(path)
        pages = []
        for page in reader.pages[:50]:  # Макс 50 страниц
            text = page.extract_text()
            if text:
                pages.append(text)
        return _truncate("\n\n".join(pages))
    except ImportError:
        pass

    logger.error("No PDF parser available")
    return None


def _truncate(text: str) -> str:
    """Обрезает текст до максимальной длины."""
    if len(text) > MAX_FILE_TEXT_LENGTH:
        return text[:MAX_FILE_TEXT_LENGTH] + "\n\n... (текст обрезан, файл слишком большой)"
    return text
